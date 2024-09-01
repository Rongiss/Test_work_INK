# для выполнения скрипта необходима установка openpyxl в виртульное окружение
# pip install openpyxl
# либо
# pip3  -install openpyxl

from openpyxl import load_workbook
import os
import warnings
warnings.filterwarnings("ignore", message="Data Validation extension is not supported")


def main():
    """
    Главная функеция
    """
    file_in_dir = find_file()
    work_type = function_column_value_counting('C', file=file_in_dir)
    work_ed = function_column_value_counting('E', file_in_dir)
    write_date(work_type=work_type, work_ed=work_ed, file=file_in_dir)


def write_date(work_type: list, work_ed: list, file):
    """
    Функция получает на вход два списка, вносит данные из списков в указанные слоббцы
    Сохраняет отдельную копию файла м изменнеными данными
    :param work_type:
    :param work_ed:
    :return: None
    """
    wb = load_workbook(file)
    ws = wb.active
    for i in range(len(work_ed)):
        ws['C' + str(i + 3)] = work_type[i].strip()
        ws['E' + str(i + 3)] = work_ed[i]
    print(f'Запись данных в "new_{file}"')
    wb.save('new_' + file)
    print('Запись завершена')


def function_column_value_counting(colom: str, file) -> list:
    """
    Функция на вход получает имя столбца и возвращает список данных находящихся в столбце
    :param colom:
    :return: list
    """
    print(f'Загрузка данных из Справочника столбец "{colom}"')

    # заливаем файл
    wb = load_workbook(file)

    # выбираем лист в таблице (в данном случае это СПравочник)
    ws = wb['Справочник']

    # указываем № столбца
    max_row = ws.max_row
    l = []

    for row in range(1, max_row + 1):
        cell = ws['{}{}'.format(colom, row)]
        if cell.value != None:
            l.append(cell.value)
        else:
            break
    print(f'Загрузка из Справочника столбец "{colom}" - завершена')
    return l[1:]


def find_file() -> str:
    """
    функция находит файл с расширением '.xlsx' возвращает его навание
    :return: str
    """
    for i in os.listdir():
        if '.xlsx' == i[-5:]:
            file_name = i
            break
    return file_name


if __name__ == '__main__':
    main()
